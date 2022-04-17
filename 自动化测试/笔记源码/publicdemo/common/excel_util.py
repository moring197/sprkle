# -*- coding:utf-8 -*-
"""
@Time 2021/3/6
@auth 码尚学院_百里老师
@Email 198977131@qq.com
@Content excel_util.py
"""


import os
import openpyxl

class ExcelUtil:

    #获得项目路径
    def get_object_path(self):
        return os.path.abspath(os.path.dirname(__file__)).split("common")[0]

    def read_excel(self):
        #openpyxl,xlrd
        #加载excel工作簿
        wb = openpyxl.load_workbook(self.get_object_path()+"data/login_data.xlsx")
        #获得sheet对象
        sheet = wb['login']
        #获得excel的行数和列数
        print(sheet.max_row,sheet.max_column)
        #循环
        all_list = []
        for rows in range(2,sheet.max_row+1):
            temp_list = []
            for cols in range(1,sheet.max_column+1):
                temp_list.append(sheet.cell(rows,cols).value)
            all_list.append(temp_list)
        return all_list

#[1,"admin","admin123"],[2,"admin","admin"],[3,"admin123","admin123"]
#[[1, 'admin', 'admin123'], [2, 'admin', 'admin'], [3, 'admin123', 'admin123']]



if __name__ == '__main__':
    ExcelUtil().read_excel()
